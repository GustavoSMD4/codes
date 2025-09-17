import base64
import csv
from io import BytesIO, StringIO
import io
import random
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from api.code.conexao.codeConexao import CodeConexao
from api.code.models.codeModel import Code
from cache.cache import Cache


class CodeService:
    def __init__(self, conexao, cache):
        self.__conexao: CodeConexao = conexao
        self.cache: Cache = cache
        
    @classmethod
    def newService(cls, cache = None):
        return cls(CodeConexao.newConexao(), cache)
    
    def getStats(self):
        return self.__conexao.stats()
    
    def getCodes(self):
        codes = self.__conexao.getCodes()
        self.cache.setCodes(codes)
        return codes
    
    def getCodesInterval(self, start, end):
        if start > end:
            raise Exception("Start needs to be lower than end.")
        
        return self.__conexao.getCodesInterval(start, end)
    
    def getCodesLimitFirst(self, req: dict):
        limit = req.get("limit")
        
        if not isinstance(limit, int):
            limit = int(limit)
            
        if limit > 10000 or limit <= 0:
            raise Exception("limit needs to be between 1 and 10000")
        
        return self.__conexao.getCodesLimit(limit)
    
    def getCodesLimitLast(self, req: dict):
        limit = req.get("limit")
        
        if not isinstance(limit, int):
            limit = int(limit)
            
        if limit > 10000 or limit <= 0:
            raise Exception("limit needs to be between 1 and 10000")
        
        return self.__conexao.getCodesLimit(limit, False)
         
    def verifyCodeSecurity(self, req: dict):        
        code: str = self.__validateCode(req.get("code"))
        
        codeDb = self.__conexao.getCodeByCode(code)
        
        return f"Your code is the number {codeDb.getId()} most used out of 10000."
    
    def suggestCode(self, req: dict):
        safe = req.get("safe")
        
        if safe == "True":
            safe = True
        elif safe == "False":
            safe = False
        else:
            raise Exception("param safe needs to be True or False.")
        
        codes = self.__conexao.suggestCode(safe)
        return Code.newModel(random.choice(codes))
    
    def gerarPlanilha(self, digits = None):
        if not isinstance(digits, int):
            digits = int(digits)
        
        if digits is None:
            digits = 3
            
        if digits not in [1, 2, 3]:
            raise Exception("digits must be 1, 2 or 3")
            
        codes = self.__conexao.getCodes()
        vistos = set()
        resultado = []

        if digits >= 4:
            raise Exception("To download the full db, use /codes/csv")
            
        else:
            for code in codes:
                prefixo = code.getCode()[:digits]
                if prefixo not in vistos:
                    vistos.add(prefixo)
                    resultado.append(prefixo)
                
        wb = Workbook()
        ws = wb.active
        ws.title = "codigos"

        ws.append(["codigo", "feito"])

        checkValidation = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
        ws.add_data_validation(checkValidation)

        for index, codigo in enumerate(resultado, start=2):
            ws[f"A{index}"] = codigo
            cellCheck = f"B{index}"
            checkValidation.add(ws[cellCheck])
            ws[cellCheck] = "FALSE"
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def gerarCSVCompleto(self):
        codes = self.__conexao.getCodes()
        registros = [code.getDict() for code in codes]

        buffer_str = StringIO()
        writer = csv.DictWriter(buffer_str, fieldnames=registros[0].keys(), delimiter=";")
        writer.writeheader()
        writer.writerows(registros)

        output = BytesIO()
        output.write(buffer_str.getvalue().encode("utf-8"))
        output.seek(0)
        return output
       
    def generateHeatmapData(self, start, end=None):
        freq = np.zeros((10, 4), dtype=int)
        
        if end is None:
            codes = self.getCodesLimitFirst({"limit": start})
            titleInfo = f'primeiros {start} códigos'
        else:
            codes = self.getCodesInterval(start, end)
            titleInfo = f'códigos de {start} até {end}'

        totalCounts = 0
        maxCode = None
        maxQtde = 0

        for code in codes:
            codeStr = code.getCode()
            qtde = code.getQtde()
            totalCounts += qtde

            if qtde > maxQtde:
                maxQtde = qtde
                maxCode = codeStr

            for pos, char in enumerate(codeStr):
                digit = int(char)
                freq[digit, pos] += qtde

        percent = np.zeros_like(freq, dtype=float)
        for pos in range(4):
            colSum = freq[:, pos].sum()
            if colSum > 0:
                percent[:, pos] = freq[:, pos] / colSum * 100

        mostCommonDigits = []
        for pos in range(4):
            col = freq[:, pos]
            mostCommonDigit = int(np.argmax(col))
            mostCommonDigits.append(mostCommonDigit)

        totalOccurrencesAll = sum(code.getQtde() or 1 for code in self.__conexao.getCodes())
        passwordPercent = (totalCounts / totalOccurrencesAll) * 100

        fig, ax = plt.subplots(figsize=(6, 4))
        cax = ax.imshow(freq, cmap='YlGnBu', aspect='auto')
        ax.set_xticks(np.arange(4))
        ax.set_yticks(np.arange(10))
        ax.set_xticklabels(['1ª posição', '2ª posição', '3ª posição', '4ª posição'])
        ax.set_yticklabels([str(i) for i in range(10)])

        for i in range(10):
            for j in range(4):
                ax.text(j, i, freq[i, j], ha='center', va='center', color='black')

        ax.set_xlabel('Posição do Dígito')
        ax.set_ylabel('Dígito')
        ax.set_title(f'Frequência dos Dígitos ({titleInfo})')
        fig.colorbar(cax, ax=ax, orientation='vertical', label='Frequência')

        buf = io.BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format='png')
        buf.seek(0)
        imgBase64 = base64.b64encode(buf.read()).decode('utf-8')
        plt.close(fig)

        return {
            "imageBase64": imgBase64,
            "frequency": freq.tolist(),
            "percent": percent.tolist(),
            "totalCodes": len(codes),
            "totalOccurrences": totalCounts,
            "mostFrequentCode": maxCode,
            "mostFrequentQtde": maxQtde,
            "mostCommonDigits": mostCommonDigits,
            "passwordPercent": passwordPercent
        }
        
    def simulateBruteForce(self, req: dict):
        code: str = self.__validateCode(req.get("code"))
    
    @staticmethod
    def __validateCode(code):
        if not isinstance(code, str):
            code = str(code)
        
        if code is None:
            raise Exception("code empty")
        
        if len(code) != 4 or not code.isdigit():
            raise Exception("code needs to have 4 digits")
        
        return code
    
        
        
        

